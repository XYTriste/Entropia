"""
全量数据级联导入服务

支持单Excel多Sheet级联导入：
专业 → 教师 → 教室 → 班级 → 课程 → 课程-班级关联 → 学生
"""

import io
from typing import Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.import_service import (
    ImportErrorReport,
    _import_majors_from_rows,
    _import_teachers_from_rows,
    _import_classrooms_from_rows,
    _import_classes_from_rows,
    _import_courses_from_rows,
    _import_course_classes_from_rows,
    _import_students_from_rows,
    TEMPLATE_CONFIG,
)


# Sheet 导入顺序定义（按依赖拓扑排序）
SHEET_IMPORT_ORDER = [
    ("majors", "专业", _import_majors_from_rows),
    ("teachers", "教师", _import_teachers_from_rows),
    ("classrooms", "教室", _import_classrooms_from_rows),
    ("classes", "班级", _import_classes_from_rows),
    ("courses", "课程", _import_courses_from_rows),
    ("course-classes", "课程班级关联", _import_course_classes_from_rows),
    ("students", "学生", _import_students_from_rows),
]


def _parse_sheet(ws) -> list[dict]:
    """将单个Worksheet解析为字典列表"""
    raw_headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    headers = [h.lstrip("*") for h in raw_headers]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if header:
                row_dict[header] = row[i] if i < len(row) else None
        rows.append(row_dict)
    return rows


def _normalize_rows(rows: list[dict]) -> list[dict]:
    """统一将值转为字符串"""
    normalized = []
    for row in rows:
        norm = {}
        for k, v in row.items():
            if v is None:
                norm[k] = ""
            elif isinstance(v, bool):
                norm[k] = "true" if v else "false"
            else:
                norm[k] = str(v)
        normalized.append(norm)
    return normalized


async def import_all_in_one(db: AsyncSession, file_bytes: bytes) -> dict[str, Any]:
    """全量数据级联导入

    解析Excel中多个Sheet，按依赖顺序依次导入。
    返回各Sheet的导入报告汇总。
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
    except Exception as e:
        return {
            "success": False,
            "overall_summary": f"Excel 解析失败: {e}",
            "sheets": [],
        }

    # 获取所有Sheet名（小写映射）
    sheet_name_map = {name.lower(): name for name in wb.sheetnames}

    overall_reports = []
    overall_success = True
    total_imported = 0
    total_errors = 0

    for entity_key, entity_label, import_func in SHEET_IMPORT_ORDER:
        # 查找对应Sheet（支持中英文Sheet名）
        sheet_name = None
        for key in (entity_key, entity_key.replace("-", ""), entity_label):
            if key in sheet_name_map:
                sheet_name = sheet_name_map[key]
                break

        if not sheet_name:
            # 该Sheet不存在，跳过（不报错，因为用户可能不需要导入某类数据）
            continue

        ws = wb[sheet_name]
        rows = _parse_sheet(ws)

        if not rows:
            continue

        rows = _normalize_rows(rows)
        report = await import_func(db, rows)

        sheet_result = {
            "sheet_name": sheet_name,
            "entity": entity_key,
            "label": entity_label,
            "success": len(report.errors) == 0,
            "success_count": report.success_count,
            "error_count": report.error_count,
            "errors": report.errors,
            "warnings": report.warnings,
        }
        overall_reports.append(sheet_result)
        total_imported += report.success_count
        total_errors += report.error_count
        if report.errors:
            overall_success = False

    summary = f"共导入 {len(overall_reports)} 个Sheet"
    if total_imported > 0:
        summary += f"，成功 {total_imported} 条"
    if total_errors > 0:
        summary += f"，失败 {total_errors} 条"

    return {
        "success": overall_success and total_errors == 0,
        "overall_summary": summary,
        "sheets": overall_reports,
    }


def generate_all_in_one_template() -> bytes:
    """生成全量数据导入模板（多Sheet）"""
    wb = Workbook()
    # 删除默认Sheet，后续按需创建
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    example_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")

    for entity_key, entity_label, _ in SHEET_IMPORT_ORDER:
        if entity_key not in TEMPLATE_CONFIG:
            continue
        config = TEMPLATE_CONFIG[entity_key]
        columns = config["columns"]

        ws = wb.create_sheet(title=entity_label)

        # 表头
        for col_idx, col in enumerate(columns, 1):
            header_text = f"*{col['key']}" if col.get("required") else col["key"]
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 示例数据
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col.get("example", ""))
            cell.fill = example_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 自动调整列宽
        for col_idx, col in enumerate(columns, 1):
            max_len = max(
                len(str(col["key"])) + 2,
                len(str(col.get("desc", ""))),
                len(str(col.get("example", ""))),
                12,
            )
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
