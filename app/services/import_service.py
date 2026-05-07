"""
考试排考系统 - CSV 导入服务

提供各类数据的 CSV 导入功能：
- 教师、教室、学生、课程、课程-班级关联
- 数据校验（学号唯一、班级唯一性、外键引用有效性）
- 兼职教师总场次容量≥60人次的警告
- 错误报告生成
- 事务导入（全部成功或全部失败）
"""

import csv
import io
from typing import Any, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.class_ import Class
from app.models.classroom import Classroom
from app.models.course import Course
from app.models.course_class import CourseClass
from app.models.major import Major
from app.models.student import Student
from app.models.teacher import Teacher
from app.models.time_slot import TimeSlot
from app.utils.validators import (
    validate_csv_classroom_row,
    validate_csv_course_class_row,
    validate_csv_course_row,
    validate_csv_student_row,
    validate_csv_teacher_row,
    validate_student_no,
)


class ImportErrorReport:
    """导入错误报告"""

    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.success_count: int = 0
        self.error_count: int = 0

    def add_error(self, msg: str) -> None:
        self.errors.append(msg)
        self.error_count += 1

    def add_warning(self, msg: str) -> None:
        self.warnings.append(msg)

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": len(self.errors) == 0,
            "success_count": self.success_count,
            "error_count": self.error_count,
            "errors": self.errors,
            "warnings": self.warnings,
        }


# ============================================================
# 教师导入
# ============================================================


async def _import_teachers_from_rows(db: AsyncSession, rows: list[dict]) -> ImportErrorReport:
    """从字典列表导入教师（供 CSV/Excel 复用）"""
    report = ImportErrorReport()
    required = {"name", "teacher_type", "max_slots"}
    if not required.issubset(set(rows[0].keys())):
        missing = required - set(rows[0].keys())
        report.add_error(f"缺少必要列: {missing}")
        return report

    teachers_to_create = []
    for line_no, row in enumerate(rows, 2):
        valid, errors = validate_csv_teacher_row(row, line_no)
        if not valid:
            for err in errors:
                report.add_error(err)
            continue
        teacher = Teacher(
            name=str(row["name"]).strip(),
            teacher_type=str(row["teacher_type"]).strip(),
            max_slots=int(row["max_slots"]),
            current_slots=0,
            is_active=True,
        )
        teachers_to_create.append(teacher)

    part_time_total = sum(t.max_slots for t in teachers_to_create if t.teacher_type == "part_time")
    if part_time_total > 0 and part_time_total < 60:
        report.add_warning(f"兼职教师总场次容量为{part_time_total}，建议≥60人次")

    if report.errors:
        return report

    db.add_all(teachers_to_create)
    await db.flush()
    report.success_count = len(teachers_to_create)
    return report


async def import_teachers_csv(
    db: AsyncSession, csv_content: str
) -> ImportErrorReport:
    """CSV 导入教师

    CSV 格式: name,teacher_type,max_slots
    """
    report = ImportErrorReport()
    reader = csv.DictReader(io.StringIO(csv_content))
    if not reader.fieldnames:
        report.add_error("CSV 格式错误: 缺少表头")
        return report
    rows = list(reader)
    if not rows:
        report.add_error("CSV 文件为空")
        return report
    return await _import_teachers_from_rows(db, rows)


# ============================================================
# 教室导入
# ============================================================


async def _import_classrooms_from_rows(db: AsyncSession, rows: list[dict]) -> ImportErrorReport:
    """从字典列表导入教室（供 CSV/Excel 复用）"""
    report = ImportErrorReport()
    required = {"name", "capacity", "room_type"}
    if not required.issubset(set(rows[0].keys())):
        missing = required - set(rows[0].keys())
        report.add_error(f"缺少必要列: {missing}")
        return report

    classrooms_to_create = []
    for line_no, row in enumerate(rows, 2):
        valid, errors = validate_csv_classroom_row(row, line_no)
        if not valid:
            for err in errors:
                report.add_error(err)
            continue

        classroom = Classroom(
            name=str(row["name"]).strip(),
            capacity=int(row["capacity"]),
            room_type=str(row["room_type"]).strip(),
            building=str(row.get("building", "")).strip(),
            floor=int(row.get("floor", 1) or 1),
            is_active=True,
        )
        classrooms_to_create.append(classroom)

    if report.errors:
        return report

    db.add_all(classrooms_to_create)
    await db.flush()
    report.success_count = len(classrooms_to_create)
    return report


async def import_classrooms_csv(
    db: AsyncSession, csv_content: str
) -> ImportErrorReport:
    """CSV 导入教室

    CSV 格式: name,capacity,room_type,building,floor
    """
    report = ImportErrorReport()
    reader = csv.DictReader(io.StringIO(csv_content))
    if not reader.fieldnames:
        report.add_error("CSV 格式错误: 缺少表头")
        return report
    rows = list(reader)
    if not rows:
        report.add_error("CSV 文件为空")
        return report
    return await _import_classrooms_from_rows(db, rows)


# ============================================================
# 学生导入
# ============================================================


async def _import_students_from_rows(db: AsyncSession, rows: list[dict]) -> ImportErrorReport:
    """从字典列表导入学生（供 CSV/Excel 复用）"""
    report = ImportErrorReport()
    required = {"student_no", "name", "class_name", "grade"}
    if not required.issubset(set(rows[0].keys())):
        missing = required - set(rows[0].keys())
        report.add_error(f"缺少必要列: {missing}")
        return report

    # 预加载所有班级 (name, grade) -> Class
    result = await db.execute(select(Class))
    all_classes: dict[tuple[str, int], int] = {}
    for c in result.scalars().all():
        all_classes[(c.name, c.grade)] = c.id

    # 预加载已有学号
    result = await db.execute(select(Student.student_no))
    existing_nos = set(result.scalars().all())

    students_to_create: list[Student] = []
    seen_nos: set[str] = set()

    for line_no, row in enumerate(rows, 2):
        valid, errors = validate_csv_student_row(row, line_no)
        if not valid:
            for err in errors:
                report.add_error(err)
            continue

        student_no = str(row["student_no"]).strip()
        class_name = str(row["class_name"]).strip()
        grade = int(row["grade"])

        # 学号唯一性校验
        if student_no in existing_nos or student_no in seen_nos:
            report.add_error(f"第{line_no}行: 学号 '{student_no}' 已存在")
            continue
        seen_nos.add(student_no)

        # 查找班级ID
        class_key = (class_name, grade)
        if class_key not in all_classes:
            report.add_error(
                f"第{line_no}行: 班级 '{class_name}'({grade}级) 不存在，请先导入班级数据"
            )
            continue

        student = Student(
            student_no=student_no,
            name=str(row["name"]).strip(),
            class_id=all_classes[class_key],
        )
        students_to_create.append(student)

    if report.errors:
        return report

    db.add_all(students_to_create)
    await db.flush()
    report.success_count = len(students_to_create)
    return report


async def import_students_csv(
    db: AsyncSession, csv_content: str
) -> ImportErrorReport:
    """CSV 导入学生

    CSV 格式: student_no,name,class_name,grade,major_name(可选)
    根据 class_name + grade 查找班级ID
    """
    report = ImportErrorReport()
    reader = csv.DictReader(io.StringIO(csv_content))
    if not reader.fieldnames:
        report.add_error("CSV 格式错误: 缺少表头")
        return report
    rows = list(reader)
    if not rows:
        report.add_error("CSV 文件为空")
        return report
    return await _import_students_from_rows(db, rows)


# ============================================================
# 课程导入
# ============================================================


async def _import_courses_from_rows(db: AsyncSession, rows: list[dict]) -> ImportErrorReport:
    """从字典列表导入课程（供 CSV/Excel 复用）"""
    report = ImportErrorReport()
    required = {"name", "course_type", "needs_ab"}
    if not required.issubset(set(rows[0].keys())):
        missing = required - set(rows[0].keys())
        report.add_error(f"缺少必要列: {missing}")
        return report

    courses_to_create = []
    for line_no, row in enumerate(rows, 2):
        valid, errors = validate_csv_course_row(row, line_no)
        if not valid:
            for err in errors:
                report.add_error(err)
            continue

        needs_ab = str(row["needs_ab"]).strip().lower() in ("true", "1", "yes")
        course_type = str(row["course_type"]).strip()

        # 公共课必须指定时段
        dept_date = row.get("dept_assigned_date")
        dept_slot = row.get("dept_assigned_time_slot_id")
        if course_type == "public":
            if not dept_date or str(dept_date).strip() == "":
                report.add_error(f"第{line_no}行: 公共课必须指定 dept_assigned_date")
                continue
            if not dept_slot or str(dept_slot).strip() == "":
                report.add_error(f"第{line_no}行: 公共课必须指定 dept_assigned_time_slot_id")
                continue

        course = Course(
            name=str(row["name"]).strip(),
            course_type=course_type,
            needs_ab=needs_ab,
            dept_assigned_date=int(dept_date) if dept_date and str(dept_date).strip() else None,
            dept_assigned_time_slot_id=int(dept_slot) if dept_slot and str(dept_slot).strip() else None,
            is_active=True,
        )
        courses_to_create.append(course)

    if report.errors:
        return report

    db.add_all(courses_to_create)
    await db.flush()
    report.success_count = len(courses_to_create)
    return report


async def import_courses_csv(
    db: AsyncSession, csv_content: str
) -> ImportErrorReport:
    """CSV 导入课程

    CSV 格式: name,course_type,needs_ab,dept_assigned_date(可选),dept_assigned_time_slot_id(可选)
    """
    report = ImportErrorReport()
    reader = csv.DictReader(io.StringIO(csv_content))
    if not reader.fieldnames:
        report.add_error("CSV 格式错误: 缺少表头")
        return report
    rows = list(reader)
    if not rows:
        report.add_error("CSV 文件为空")
        return report
    return await _import_courses_from_rows(db, rows)


# ============================================================
# 课程-班级关联导入
# ============================================================


async def _import_course_classes_from_rows(db: AsyncSession, rows: list[dict]) -> ImportErrorReport:
    """从字典列表导入课程-班级关联（供 CSV/Excel 复用）"""
    report = ImportErrorReport()
    required = {"course_name", "class_name", "grade"}
    if not required.issubset(set(rows[0].keys())):
        missing = required - set(rows[0].keys())
        report.add_error(f"缺少必要列: {missing}")
        return report

    # 预加载课程名 -> ID
    result = await db.execute(select(Course))
    course_map: dict[str, int] = {}
    for c in result.scalars().all():
        course_map[c.name] = c.id

    # 预加载班级 (name, grade) -> ID
    result = await db.execute(select(Class))
    class_map: dict[tuple[str, int], int] = {}
    for c in result.scalars().all():
        class_map[(c.name, c.grade)] = c.id

    links_to_create: list[CourseClass] = []
    seen_links: set[tuple[int, int, int]] = set()

    for line_no, row in enumerate(rows, 2):
        valid, errors = validate_csv_course_class_row(row, line_no)
        if not valid:
            for err in errors:
                report.add_error(err)
            continue

        course_name = str(row["course_name"]).strip()
        class_name = str(row["class_name"]).strip()
        grade = int(row["grade"])

        if course_name not in course_map:
            report.add_error(f"第{line_no}行: 课程 '{course_name}' 不存在")
            continue

        class_key = (class_name, grade)
        if class_key not in class_map:
            report.add_error(f"第{line_no}行: 班级 '{class_name}'({grade}级) 不存在")
            continue

        course_id = course_map[course_name]
        class_id = class_map[class_key]
        link_key = (course_id, class_id, grade)

        if link_key in seen_links:
            report.add_warning(f"第{line_no}行: 关联重复，已跳过")
            continue
        seen_links.add(link_key)

        links_to_create.append(
            CourseClass(course_id=course_id, class_id=class_id, grade=grade)
        )

    if report.errors:
        return report

    db.add_all(links_to_create)
    await db.flush()
    report.success_count = len(links_to_create)
    return report


async def import_course_classes_csv(
    db: AsyncSession, csv_content: str
) -> ImportErrorReport:
    """CSV 导入课程-班级关联

    CSV 格式: course_name,class_name,grade
    """
    report = ImportErrorReport()
    reader = csv.DictReader(io.StringIO(csv_content))
    if not reader.fieldnames:
        report.add_error("CSV 格式错误: 缺少表头")
        return report
    rows = list(reader)
    if not rows:
        report.add_error("CSV 文件为空")
        return report
    return await _import_course_classes_from_rows(db, rows)


# ============================================================
# 通用数据校验
# ============================================================


async def validate_all_data(db: AsyncSession) -> dict[str, Any]:
    """校验全部数据的完整性与一致性

    返回校验报告，包含错误和警告。
    """
    errors: list[str] = []
    warnings: list[str] = []

    # 1. 检查是否有教师
    result = await db.execute(select(Teacher))
    teachers = result.scalars().all()
    if not teachers:
        errors.append("教师表为空，请先导入教师数据")
    else:
        # 检查兼职教师总场次
        part_time_total = sum(t.max_slots for t in teachers if t.teacher_type == "part_time")
        if part_time_total < 60:
            warnings.append(f"兼职教师总场次容量为{part_time_total}，建议≥60人次")

    # 2. 检查是否有教室
    result = await db.execute(select(Classroom))
    classrooms = result.scalars().all()
    if not classrooms:
        errors.append("教室表为空，请先导入教室数据")
    else:
        active_capacity = sum(c.capacity for c in classrooms if c.is_active)
        if active_capacity == 0:
            errors.append("没有启用的教室")

    # 3. 检查是否有班级
    result = await db.execute(select(Class))
    classes = result.scalars().all()
    if not classes:
        errors.append("班级表为空，请先导入班级数据")

    # 4. 检查是否有学生
    result = await db.execute(select(Student))
    students = result.scalars().all()
    if not students:
        warnings.append("学生表为空，如需要学生明细请导入")

    # 5. 检查是否有课程
    result = await db.execute(select(Course))
    courses = result.scalars().all()
    if not courses:
        errors.append("课程表为空，请先导入课程数据")
    else:
        # 检查课程是否有关联班级
        for c in courses:
            if not c.class_links:
                warnings.append(f"课程 '{c.name}' 没有关联任何班级")

    # 6. 检查时段表是否已初始化
    result = await db.execute(select(TimeSlot))
    from app.models.time_slot import TimeSlot
    time_slots = result.scalars().all()
    if not time_slots:
        warnings.append("时段表为空，系统将自动初始化默认时段")

    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
    }


# ============================================================
# 班级导入
# ============================================================


async def _import_classes_from_rows(db: AsyncSession, rows: list[dict]) -> ImportErrorReport:
    """从字典列表导入班级（支持 major_id 或 major_name 引用专业）"""
    report = ImportErrorReport()
    has_major_id = "major_id" in rows[0]
    has_major_name = "major_name" in rows[0]
    if not has_major_id and not has_major_name:
        report.add_error("缺少必要列: major_id 或 major_name 至少提供一个")
        return report
    if "name" not in rows[0] or "grade" not in rows[0]:
        missing = []
        if "name" not in rows[0]:
            missing.append("name")
        if "grade" not in rows[0]:
            missing.append("grade")
        report.add_error(f"缺少必要列: {missing}")
        return report

    # 预加载所有专业 (id -> Major, name -> id)
    result = await db.execute(select(Major))
    major_by_id = {}
    major_by_name = {}
    for m in result.scalars().all():
        major_by_id[m.id] = m
        major_by_name[m.name] = m.id

    classes_to_create = []
    for line_no, row in enumerate(rows, 2):
        name = str(row.get("name", "")).strip()
        if not name:
            report.add_error(f"第{line_no}行: 班级名称为空")
            continue

        # 解析 major_id
        major_id = None
        if has_major_id and row.get("major_id", "").strip():
            try:
                major_id = int(row["major_id"])
            except (ValueError, TypeError):
                report.add_error(f"第{line_no}行: major_id 必须是整数")
                continue
        elif has_major_name:
            major_name = str(row.get("major_name", "")).strip()
            if major_name:
                if major_name in major_by_name:
                    major_id = major_by_name[major_name]
                else:
                    report.add_error(f"第{line_no}行: 专业名称 '{major_name}' 不存在")
                    continue

        if major_id is None:
            report.add_error(f"第{line_no}行: 未提供有效的 major_id 或 major_name")
            continue
        if major_id not in major_by_id:
            report.add_error(f"第{line_no}行: 专业ID {major_id} 不存在")
            continue

        try:
            grade = int(row["grade"])
            if grade < 1 or grade > 4:
                report.add_error(f"第{line_no}行: grade 范围应在1-4之间")
                continue
        except (ValueError, TypeError):
            report.add_error(f"第{line_no}行: grade 必须是整数")
            continue
        try:
            student_count = int(row.get("student_count", 0) or 0)
        except (ValueError, TypeError):
            student_count = 0

        cls = Class(name=name, major_id=major_id, grade=grade, student_count=student_count)
        classes_to_create.append(cls)

    if report.errors:
        return report

    db.add_all(classes_to_create)
    await db.flush()
    report.success_count = len(classes_to_create)
    return report


# ============================================================
# 专业导入
# ============================================================


async def _import_majors_from_rows(db: AsyncSession, rows: list[dict]) -> ImportErrorReport:
    """从字典列表导入专业"""
    report = ImportErrorReport()
    if "name" not in rows[0]:
        report.add_error("缺少必要列: name")
        return report

    # 预加载已有专业名
    result = await db.execute(select(Major.name))
    existing_names = set(result.scalars().all())

    majors_to_create = []
    seen = set()
    for line_no, row in enumerate(rows, 2):
        name = str(row.get("name", "")).strip()
        if not name:
            report.add_error(f"第{line_no}行: 专业名称为空")
            continue
        if name in existing_names or name in seen:
            report.add_error(f"第{line_no}行: 专业 '{name}' 已存在")
            continue
        seen.add(name)
        majors_to_create.append(Major(name=name))

    if report.errors:
        return report

    db.add_all(majors_to_create)
    await db.flush()
    report.success_count = len(majors_to_create)
    return report


# ============================================================
# Excel 解析
# ============================================================


def parse_excel(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    """将 Excel 文件解析为 (headers, rows_dict_list)

    表头中的 * 前缀表示必填列，解析时自动去除。
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    raw_headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    headers = [h.lstrip("*") for h in raw_headers]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if header:
                row_dict[header] = row[i] if i < len(row) else None
        rows.append(row_dict)
    return headers, rows


# ============================================================
# 模板配置和生成
# ============================================================


TEMPLATE_CONFIG = {
    "teachers": {
        "sheet_name": "教师导入模板",
        "columns": [
            {"key": "name", "required": True, "desc": "教师姓名，字符串，长度≤50", "example": "张三"},
            {"key": "teacher_type", "required": True, "desc": "教师类型: full_time(专任) 或 part_time(兼职)", "example": "full_time"},
            {"key": "max_slots", "required": True, "desc": "最大监考场次上限，整数≥0", "example": 6},
        ],
        "notes": [
            "教师类型说明: full_time=专任教师，part_time=兼职教师",
            "兼职教师总场次容量建议≥60人次，否则系统会发出警告",
        ],
    },
    "classrooms": {
        "sheet_name": "教室导入模板",
        "columns": [
            {"key": "name", "required": True, "desc": "教室名称，全局唯一，如 A101、J101", "example": "A101"},
            {"key": "capacity", "required": True, "desc": "容纳人数，整数，如 50", "example": 50},
            {"key": "room_type", "required": True, "desc": "教室类型: regular(普通教室) 或 lecture(阶梯教室)", "example": "regular"},
            {"key": "building", "required": False, "desc": "所在教学楼名称，如 A教学楼", "example": "A教学楼"},
            {"key": "floor", "required": False, "desc": "所在楼层，整数，默认1", "example": 1},
        ],
        "notes": [],
    },
    "students": {
        "sheet_name": "学生导入模板",
        "columns": [
            {"key": "student_no", "required": True, "desc": "学号，全局唯一，如 STU000001", "example": "STU000001"},
            {"key": "name", "required": True, "desc": "学生姓名", "example": "学生1"},
            {"key": "class_name", "required": True, "desc": "班级名称，必须与系统中已存在的班级名称完全一致", "example": "计算机科学与技术大一1班"},
            {"key": "grade", "required": True, "desc": "年级: 1=大一, 2=大二, 3=大三, 4=大四", "example": 1},
        ],
        "notes": [
            "导入前请确保对应班级已存在于系统中",
            "学号全局唯一，重复导入会报错",
        ],
    },
    "courses": {
        "sheet_name": "课程导入模板",
        "columns": [
            {"key": "name", "required": True, "desc": "课程名称，如 高等数学A", "example": "高等数学A"},
            {"key": "course_type", "required": True, "desc": "课程类型: public(公共课/全校统一) 或 major(专业课/学院自行)", "example": "public"},
            {"key": "needs_ab", "required": True, "desc": "是否需要分AB卷考试: true/false/1/0", "example": "true"},
            {"key": "dept_assigned_date", "required": False, "desc": "公共课已分配日期: 1=周一, 2=周二, ..., 5=周五；公共课必填，专业课留空", "example": 1},
            {"key": "dept_assigned_time_slot_id", "required": False, "desc": "公共课已分配时段ID，对应时段表主键；公共课必填，专业课留空", "example": 1},
        ],
        "notes": [
            "公共课(course_type=public)必须填写 dept_assigned_date 和 dept_assigned_time_slot_id",
            "专业课(course_type=major)不需要填写后两列",
            "needs_ab=true 表示同课程分两场考试(A卷和B卷)",
            "时段ID对照: 1=周一T1, 2=周一T2, 3=周一T3, 4=周一T4, 5=周二T1, 6=周二T2, 7=周二T3, 8=周二T4,",
            "             9=周三T1, 10=周三T2, 11=周三T3, 12=周三T4, 13=周四T1, 14=周四T2, 15=周四T3, 16=周四T4,",
            "             17=周五T1, 18=周五T2, 19=周五T3, 20=周五T4",
            "AB卷公共课: A卷使用指定时段, B卷自动分配为下一个连续时段(T1→T2, T3→T4)",
        ],
    },
    "classes": {
        "sheet_name": "班级导入模板",
        "columns": [
            {"key": "name", "required": True, "desc": "班级名称，如 计算机科学与技术大一1班", "example": "计算机科学与技术大一1班"},
            {"key": "major_name", "required": True, "desc": "所属专业名称（与专业Sheet中的name对应）", "example": "计算机科学"},
            {"key": "grade", "required": True, "desc": "年级: 1=大一, 2=大二, 3=大三, 4=大四", "example": 1},
            {"key": "student_count", "required": False, "desc": "学生人数，整数，默认0", "example": 30},
        ],
        "notes": [
            "支持通过 major_name 引用专业（级联导入推荐）",
            "如需单独导入，也可使用 major_id 列替代 major_name",
            "同一专业下，(name, grade) 联合唯一",
        ],
    },
    "majors": {
        "sheet_name": "专业导入模板",
        "columns": [
            {"key": "name", "required": True, "desc": "专业名称，全局唯一，如 计算机科学与技术", "example": "计算机科学与技术"},
        ],
        "notes": [],
    },
    "course-classes": {
        "sheet_name": "课程班级关联导入模板",
        "columns": [
            {"key": "course_name", "required": True, "desc": "课程名称，必须与系统中已存在的课程名称完全一致", "example": "高等数学A"},
            {"key": "class_name", "required": True, "desc": "班级名称，必须与系统中已存在的班级名称完全一致", "example": "计算机科学与技术大一1班"},
            {"key": "grade", "required": True, "desc": "年级: 1=大一, 2=大二, 3=大三, 4=大四", "example": 1},
        ],
        "notes": [
            "导入前请确保对应课程和班级已存在于系统中",
            "同一课程的同一班级在同一 grade 下不能重复关联",
        ],
    },
    "time-slots": {
        "sheet_name": "考试时段导入模板",
        "columns": [
            {"key": "day_of_week", "required": True, "desc": "星期: 1=周一, 2=周二, 3=周三, 4=周四, 5=周五", "example": 1},
            {"key": "slot_code", "required": True, "desc": "时段编码: T1/T2/T3/T4", "example": "T1"},
            {"key": "start_time", "required": True, "desc": "开始时间，格式 HH:MM", "example": "08:30"},
            {"key": "end_time", "required": True, "desc": "结束时间，格式 HH:MM", "example": "10:10"},
            {"key": "is_continuous", "required": False, "desc": "是否与下一场连续: true/false", "example": "true"},
        ],
        "notes": [
            "标准时段参考:",
            "周一到周五，每天4个时段",
            "上午: T1=08:30-10:10, T2=10:20-12:00",
            "下午: T3=14:00-15:40, T4=15:50-17:30",
            "同一(day_of_week, slot_code)组合不能重复",
        ],
    },
}


def generate_excel_template(entity: str) -> bytes:
    """生成指定实体的 Excel 导入模板

    数据 Sheet:
        第1行: 列名（必填列前缀带 *）
        第2行: 示例数据
    说明 Sheet:
        列说明表格 + 注意事项
    """
    if entity not in TEMPLATE_CONFIG:
        raise ValueError(f"不支持的实体类型: {entity}")

    config = TEMPLATE_CONFIG[entity]
    columns = config["columns"]
    wb = Workbook()
    ws = wb.active
    ws.title = config["sheet_name"]

    # 样式定义
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    required_mark_font = Font(color="FF0000", bold=True, size=11)
    example_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")

    # ===== 数据 Sheet =====
    # 第1行: 表头（必填列加 * 前缀）
    for col_idx, col in enumerate(columns, 1):
        header_text = f"*{col['key']}" if col.get("required") else col["key"]
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 第2行: 示例数据（浅蓝背景标注）
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col.get("example", ""))
        cell.fill = example_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自动调整列宽
    for col_idx, col in enumerate(columns, 1):
        max_len = max(
            len(str(col["key"])) + 2,
            len(str(col.get("desc", ""))),
            len(str(col.get("example", ""))),
            12,
        )
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ===== 说明 Sheet =====
    ws_info = wb.create_sheet("填写说明")
    ws_info.cell(row=1, column=1, value=f"{config['sheet_name']} - 填写说明")
    title_cell = ws_info.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    ws_info.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # 列说明表头
    info_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    info_header_font = Font(color="FFFFFF", bold=True, size=11)
    info_headers = [("列名", 20), ("是否必填", 12), ("说明", 60), ("示例", 20)]
    for col_idx, (text, width) in enumerate(info_headers, 1):
        cell = ws_info.cell(row=3, column=col_idx, value=text)
        cell.fill = info_header_fill
        cell.font = info_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_info.column_dimensions[cell.column_letter].width = width

    # 列说明内容
    for i, col in enumerate(columns):
        r = 4 + i
        ws_info.cell(row=r, column=1, value=col["key"])
        req_cell = ws_info.cell(row=r, column=2, value="是" if col.get("required") else "否")
        if col.get("required"):
            req_cell.font = Font(color="FF0000", bold=True)
        ws_info.cell(row=r, column=3, value=col.get("desc", ""))
        ws_info.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws_info.cell(row=r, column=4, value=col.get("example", ""))

    # 额外说明
    notes = config.get("notes", [])
    row = 4 + len(columns) + 2
    ws_info.cell(row=row, column=1, value="注意事项：")
    ws_info.cell(row=row, column=1).font = Font(bold=True, size=12)
    default_notes = [
        "1. 请保留第一行表头（含 * 前缀），不要修改列名",
        "2. 示例数据可以删除，从第2行开始填写真实数据",
        "3. 带 * 号的列为必填项，留空会导致导入失败",
        "4. 保存为 .xlsx 格式后上传",
    ]
    for i, note in enumerate(default_notes + notes):
        ws_info.cell(row=row + 1 + i, column=1, value=note)
        ws_info.cell(row=row + 1 + i, column=1).alignment = Alignment(wrap_text=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ============================================================
# Excel 导入入口
# ============================================================


async def _import_time_slots_from_rows(db: AsyncSession, rows: list[dict]) -> ImportErrorReport:
    """从字典列表导入时段"""
    report = ImportErrorReport()
    required = {"day_of_week", "slot_code", "start_time", "end_time"}
    if not required.issubset(set(rows[0].keys())):
        missing = required - set(rows[0].keys())
        report.add_error(f"缺少必要列: {missing}")
        return report

    # 预加载已有时段 (day_of_week, slot_code) -> bool
    result = await db.execute(select(TimeSlot))
    existing_keys = {(ts.day_of_week, ts.slot_code) for ts in result.scalars().all()}

    slots_to_create = []
    for line_no, row in enumerate(rows, 2):
        try:
            dow = int(row["day_of_week"])
            if dow < 1 or dow > 5:
                report.add_error(f"第{line_no}行: day_of_week 必须在 1-5 之间")
                continue
        except (ValueError, TypeError):
            report.add_error(f"第{line_no}行: day_of_week 必须是整数")
            continue

        slot_code = str(row["slot_code"]).strip().upper()
        if slot_code not in {"T1", "T2", "T3", "T4"}:
            report.add_error(f"第{line_no}行: slot_code 必须是 T1/T2/T3/T4 之一")
            continue

        start_time = str(row["start_time"]).strip()
        end_time = str(row["end_time"]).strip()
        if not start_time or not end_time:
            report.add_error(f"第{line_no}行: start_time 和 end_time 不能为空")
            continue

        key = (dow, slot_code)
        if key in existing_keys:
            report.add_warning(f"第{line_no}行: 时段 ({dow}, {slot_code}) 已存在，已跳过")
            continue
        existing_keys.add(key)

        is_continuous = str(row.get("is_continuous", "true")).strip().lower() in ("true", "1", "yes")
        slots_to_create.append(TimeSlot(
            day_of_week=dow,
            slot_code=slot_code,
            start_time=start_time,
            end_time=end_time,
            is_continuous=is_continuous,
        ))

    if report.errors:
        return report

    db.add_all(slots_to_create)
    await db.flush()
    report.success_count = len(slots_to_create)
    return report


async def import_excel(db: AsyncSession, file_bytes: bytes, entity: str) -> ImportErrorReport:
    """Excel 批量导入入口"""
    try:
        headers, rows = parse_excel(file_bytes)
    except Exception as e:
        report = ImportErrorReport()
        report.add_error(f"Excel 解析失败: {e}")
        return report

    if not rows:
        report = ImportErrorReport()
        report.add_error("Excel 文件为空或没有数据行")
        return report

    # 统一将值转为字符串，与 CSV 格式保持一致
    normalized_rows = []
    for row in rows:
        normalized = {}
        for k, v in row.items():
            if v is None:
                normalized[k] = ""
            elif isinstance(v, bool):
                normalized[k] = "true" if v else "false"
            else:
                normalized[k] = str(v)
        normalized_rows.append(normalized)
    rows = normalized_rows

    if entity == "teachers":
        return await _import_teachers_from_rows(db, rows)
    elif entity == "classrooms":
        return await _import_classrooms_from_rows(db, rows)
    elif entity == "students":
        return await _import_students_from_rows(db, rows)
    elif entity == "courses":
        return await _import_courses_from_rows(db, rows)
    elif entity == "classes":
        return await _import_classes_from_rows(db, rows)
    elif entity == "majors":
        return await _import_majors_from_rows(db, rows)
    elif entity == "course-classes":
        return await _import_course_classes_from_rows(db, rows)
    elif entity == "time-slots":
        return await _import_time_slots_from_rows(db, rows)
    else:
        report = ImportErrorReport()
        report.add_error(f"不支持的实体类型: {entity}")
        return report
