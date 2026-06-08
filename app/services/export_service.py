"""
考试排考系统 - 导出服务

提供排考结果的多种格式导出：
- Excel 多 Sheet 导出 (openpyxl)
  - 总览表、教师监考表、班级通知表、考场签到表、流动监考巡查表
- JSON 格式导出
- SQL 格式导出
"""

import json
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.class_ import Class
from app.models.classroom import Classroom
from app.models.course import Course
from app.models.exam import Exam
from app.models.exam_classroom import ExamClassroom
from app.models.exam_classroom_class import ExamClassroomClass
from app.models.exam_teacher import ExamTeacher
from app.models.patrol_teacher import PatrolTeacher
from app.models.schedule_version import ScheduleVersion, ScheduleVersionStatus
from app.models.teacher import Teacher
from app.models.time_slot import TimeSlot

# 国内时区 UTC+8
CN_TZ = timezone(timedelta(hours=8))


# ============================================================
# 样式常量
# ============================================================

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DAY_NAMES = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五"}


def _set_header_style(cell):
    """设置表头样式"""
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = BORDER


def _set_cell_style(cell):
    """设置普通单元格样式"""
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER


def _auto_width(worksheet, min_width: int = 10, max_width: int = 40):
    """自动调整列宽"""
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        adjusted_width = min(max(length + 2, min_width), max_width)
        worksheet.column_dimensions[col_letter].width = adjusted_width


# ============================================================
# 数据加载
# ============================================================


async def _load_exams_with_relations(db: AsyncSession, version_id: int | None = None) -> list[Exam]:
    """加载所有考试及其关联数据

    Args:
        version_id: 如果指定，仅导出版本对应的考试数据（需是PUBLISHED状态）
    """
    query = select(Exam)

    # 如果指定版本且是 PUBLISHED，只导出该版本数据
    if version_id:
        version_result = await db.execute(
            select(ScheduleVersion).where(ScheduleVersion.id == version_id)
        )
        version = version_result.scalar_one_or_none()
        if version and version.status == ScheduleVersionStatus.PUBLISHED:
            # PUBLISHED 版本的数据在 exams 表中，status=SCHEDULED
            query = query.where(Exam.status == "scheduled")
        elif version:
            # 非 PUBLISHED 版本（draft/archived），数据在快照中
            # 临时解析快照返回虚拟数据
            return await _load_exams_from_snapshot(version.data_snapshot, db)
        else:
            return []

    query = query.options(
        selectinload(Exam.course),
        selectinload(Exam.time_slot),
        selectinload(Exam.classroom_assignments).selectinload(ExamClassroom.class_assignments).selectinload(ExamClassroomClass.class_),
        selectinload(Exam.classroom_assignments).selectinload(ExamClassroom.classroom),
        selectinload(Exam.teacher_assignments).selectinload(ExamTeacher.teacher),
    ).order_by(Exam.time_slot_id)

    result = await db.execute(query)
    return list(result.scalars().all())


async def _load_exams_from_snapshot(snapshot_data: str | None, db: AsyncSession) -> list[dict]:
    """从快照数据加载考试信息（用于非PUBLISHED版本）"""
    if not snapshot_data:
        return []

    try:
        snapshot = json.loads(snapshot_data)
    except json.JSONDecodeError:
        return []

    # 返回快照中的原始数据（不转换为 Exam 模型）
    return snapshot.get("exams", [])


async def _load_time_slots(db: AsyncSession) -> list[TimeSlot]:
    """加载所有时段"""
    result = await db.execute(
        select(TimeSlot)
        .options(selectinload(TimeSlot.patrol_teachers).selectinload(PatrolTeacher.teacher))
        .order_by(TimeSlot.id)
    )
    return list(result.scalars().all())


async def _load_teachers(db: AsyncSession) -> list[Teacher]:
    """加载所有教师"""
    result = await db.execute(select(Teacher))
    return list(result.scalars().all())


async def _load_classrooms(db: AsyncSession) -> list[Classroom]:
    """加载所有教室"""
    result = await db.execute(select(Classroom))
    return list(result.scalars().all())


async def _load_classes(db: AsyncSession) -> list[Class]:
    """加载所有班级"""
    result = await db.execute(select(Class))
    return list(result.scalars().all())


# ============================================================
# Excel 导出 - 主入口
# ============================================================


async def export_excel(db: AsyncSession, version_id: int | None = None) -> bytes:
    """导出排考结果为 Excel 文件 (多 Sheet)

    Args:
        version_id: 可选，指定版本ID，仅导出版本对应的排考数据
    """
    wb = Workbook()

    exams = await _load_exams_with_relations(db, version_id=version_id)
    time_slots = await _load_time_slots(db)
    teachers = await _load_teachers(db)
    classrooms_data = await _load_classrooms(db)
    classes_data = await _load_classes(db)

    # Sheet 1: 排考总览表
    _build_overview_sheet(wb.active, exams, time_slots, classrooms_data, teachers)
    wb.active.title = "排考总览表"

    # Sheet 2: 教师监考表
    wb.create_sheet("教师监考表")
    _build_teacher_sheet(wb["教师监考表"], exams, time_slots, teachers)

    # Sheet 3: 班级通知表
    wb.create_sheet("班级通知表")
    _build_class_notice_sheet(wb["班级通知表"], exams, time_slots, classes_data, classrooms_data)

    # Sheet 4: 考场签到表
    wb.create_sheet("考场签到表")
    _build_classroom_sign_sheet(wb["考场签到表"], exams, time_slots, classrooms_data)

    # Sheet 5: 流动监考巡查表
    wb.create_sheet("流动监考巡查表")
    _build_patrol_sheet(wb["流动监考巡查表"], time_slots, teachers)

    # Sheet 6: 教师监考场次统计表
    wb.create_sheet("教师监考场次统计表")
    _build_teacher_invigilation_stats_sheet(
        wb["教师监考场次统计表"], exams, time_slots, teachers
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


async def export_teacher_stats_excel(
    db: AsyncSession, version_id: int | None = None
) -> bytes:
    """单独导出教师监考场次统计表

    Args:
        version_id: 可选，指定版本ID，仅导出版本对应的排考数据
    """
    wb = Workbook()
    exams = await _load_exams_with_relations(db, version_id=version_id)
    time_slots = await _load_time_slots(db)
    teachers = await _load_teachers(db)

    _build_teacher_invigilation_stats_sheet(wb.active, exams, time_slots, teachers)
    wb.active.title = "教师监考场次统计表"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ---------- Sheet 1: 排考总览表 ----------


def _build_overview_sheet(ws, exams, time_slots, classrooms, teachers):
    """构建排考总览表（按考试-教室-班级粒度展开）"""
    headers = [
        "序号", "课程名称", "AB卷", "星期", "时段", "时间",
        "教室", "监考教师", "班级", "人数",
    ]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    ts_map = {ts.id: ts for ts in time_slots}
    teacher_map = {t.id: t for t in teachers}
    room_map = {r.id: r for r in classrooms}

    row_idx = 1
    seq = 0
    prev_course_label = None

    for exam in sorted(exams, key=lambda e: (e.time_slot_id or 0, e.course.name)):
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        day_str = DAY_NAMES.get(ts.day_of_week, "") if ts else ""
        time_str = f"{ts.start_time}-{ts.end_time}" if ts else ""
        label_str = exam.exam_label.value if exam.exam_label else ""
        course_label = (exam.course.name, label_str)

        # 同一课程（含AB卷标签相同）只增加一次序号
        if course_label != prev_course_label:
            seq += 1
            prev_course_label = course_label

        # 构建教室 -> 监考教师的映射
        room_teachers: dict[int, list[str]] = {}
        exam_fixed_teachers: list[str] = []
        for et in exam.teacher_assignments:
            if et.role.value == "fixed":
                t = teacher_map.get(et.teacher_id)
                if t:
                    if et.classroom_id:
                        room_teachers.setdefault(et.classroom_id, []).append(t.name)
                    else:
                        exam_fixed_teachers.append(t.name)

        # 按教室-班级展开为行
        for ec in exam.classroom_assignments:
            room = room_map.get(ec.classroom_id)
            room_name = room.name if room else f"教室{ec.classroom_id}"
            teachers_in_room = room_teachers.get(ec.classroom_id, [])
            if teachers_in_room:
                teacher_str = "、".join(teachers_in_room)
            elif exam_fixed_teachers:
                teacher_str = "、".join(exam_fixed_teachers)
            else:
                teacher_str = ""

            if ec.class_assignments:
                for ca in ec.class_assignments:
                    cls_name = ca.class_.name if hasattr(ca, "class_") and ca.class_ else ""
                    row_idx += 1
                    ws.append([
                        seq, exam.course.name, label_str, day_str,
                        ts.slot_code if ts else "", time_str, room_name,
                        teacher_str, cls_name, ca.student_count,
                    ])
                    for cell in ws[row_idx]:
                        _set_cell_style(cell)
            else:
                # 兼容：没有班级明细时，只输出教室总人数一行
                row_idx += 1
                ws.append([
                    seq, exam.course.name, label_str, day_str,
                    ts.slot_code if ts else "", time_str, room_name,
                    teacher_str, "", ec.total_students,
                ])
                for cell in ws[row_idx]:
                    _set_cell_style(cell)

    _auto_width(ws)


# ---------- Sheet 2: 教师监考表 ----------


def _build_teacher_sheet(ws, exams, time_slots, teachers):
    """构建教师监考表"""
    headers = ["教师姓名", "教师类型", "星期", "时段", "时间", "课程", "AB卷", "监考角色", "教室"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    ts_map = {ts.id: ts for ts in time_slots}
    room_map = {}
    for exam in exams:
        for ec in exam.classroom_assignments:
            room_name = ec.classroom.name if hasattr(ec, "classroom") and ec.classroom else f"教室{ec.classroom_id}"
            room_map[ec.classroom_id] = room_name

    teacher_exams: dict[int, list[dict]] = {}
    for exam in exams:
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        for et in exam.teacher_assignments:
            if et.teacher_id not in teacher_exams:
                teacher_exams[et.teacher_id] = []
            info = {
                "day": DAY_NAMES.get(ts.day_of_week, "") if ts else "",
                "slot": ts.slot_code if ts else "",
                "time": f"{ts.start_time}-{ts.end_time}" if ts else "",
                "course": exam.course.name,
                "label": exam.exam_label.value if exam.exam_label else "",
                "role": "固定监考" if et.role.value == "fixed" else "流动监考",
                "room": "",
            }
            if et.role.value == "fixed" and et.classroom_id:
                info["room"] = room_map.get(et.classroom_id, f"教室{et.classroom_id}")
            teacher_exams[et.teacher_id].append(info)

    row_idx = 1
    for teacher in sorted(teachers, key=lambda t: t.name):
        if teacher.id not in teacher_exams:
            continue
        for info in sorted(teacher_exams[teacher.id], key=lambda x: (x["day"], x["slot"])):
            row_idx += 1
            type_str = "专任" if teacher.teacher_type.value == "full_time" else "兼职"
            ws.append([
                teacher.name, type_str, info["day"], info["slot"],
                info["time"], info["course"], info["label"], info["role"], info["room"],
            ])
            for cell in ws[row_idx]:
                _set_cell_style(cell)

    _auto_width(ws)


# ---------- Sheet 3: 班级通知表 ----------


def _build_class_notice_sheet(ws, exams, time_slots, classes_data, classrooms_data):
    """构建班级通知表"""
    headers = ["班级名称", "年级", "星期", "时段", "时间", "课程", "AB卷", "考场", "考生数"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    ts_map = {ts.id: ts for ts in time_slots}
    class_map = {c.id: c for c in classes_data}
    room_map = {r.id: r for r in classrooms_data}

    row_idx = 1
    for exam in sorted(exams, key=lambda e: e.time_slot_id or 0):
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        for ec in exam.classroom_assignments:
            for ca in ec.class_assignments:
                cls = class_map.get(ca.class_id)
                if not cls:
                    continue
                room = room_map.get(ec.classroom_id)
                row_idx += 1
                ws.append([
                    cls.name, cls.grade,
                    DAY_NAMES.get(ts.day_of_week, "") if ts else "",
                    ts.slot_code if ts else "",
                    f"{ts.start_time}-{ts.end_time}" if ts else "",
                    exam.course.name,
                    exam.exam_label.value if exam.exam_label else "",
                    room.name if room else f"教室{ec.classroom_id}",
                    ca.student_count,
                ])
                for cell in ws[row_idx]:
                    _set_cell_style(cell)

    _auto_width(ws)


# ---------- Sheet 4: 考场签到表 ----------


def _build_classroom_sign_sheet(ws, exams, time_slots, classrooms_data):
    """构建考场签到表"""
    headers = ["教室名称", "星期", "时段", "时间", "课程", "AB卷", "应到人数", "监考教师"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    ts_map = {ts.id: ts for ts in time_slots}
    teacher_map = {}
    for exam in exams:
        for et in exam.teacher_assignments:
            if hasattr(et, "teacher") and et.teacher:
                teacher_map[et.teacher_id] = et.teacher.name
    room_map = {r.id: r for r in classrooms_data}

    row_idx = 1
    for exam in sorted(exams, key=lambda e: e.time_slot_id or 0):
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        for ec in exam.classroom_assignments:
            room = room_map.get(ec.classroom_id)
            fixed_t = []
            for et in exam.teacher_assignments:
                if et.role.value == "fixed" and et.classroom_id == ec.classroom_id:
                    name = teacher_map.get(et.teacher_id, f"教师{et.teacher_id}")
                    fixed_t.append(name)

            row_idx += 1
            ws.append([
                room.name if room else f"教室{ec.classroom_id}",
                DAY_NAMES.get(ts.day_of_week, "") if ts else "",
                ts.slot_code if ts else "",
                f"{ts.start_time}-{ts.end_time}" if ts else "",
                exam.course.name,
                exam.exam_label.value if exam.exam_label else "",
                ec.total_students,
                "; ".join(fixed_t),
            ])
            for cell in ws[row_idx]:
                _set_cell_style(cell)

    _auto_width(ws)


# ---------- Sheet 5: 流动监考巡查表 ----------


def _build_patrol_sheet(ws, time_slots, teachers):
    """构建流动监考巡查表"""
    headers = ["星期", "时段", "时间", "流动监考1", "流动监考2", "流动监考3"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    teacher_map = {t.id: t for t in teachers}

    row_idx = 1
    for ts in sorted(time_slots, key=lambda t: (t.day_of_week, t.slot_code)):
        patrol_names = []
        for pt in ts.patrol_teachers:
            t = teacher_map.get(pt.teacher_id)
            patrol_names.append(t.name if t else f"教师{pt.teacher_id}")
        while len(patrol_names) < 3:
            patrol_names.append("")

        row_idx += 1
        ws.append([
            DAY_NAMES.get(ts.day_of_week, ""),
            ts.slot_code,
            f"{ts.start_time}-{ts.end_time}",
            patrol_names[0], patrol_names[1], patrol_names[2],
        ])
        for cell in ws[row_idx]:
            _set_cell_style(cell)

    _auto_width(ws)


# ---------- Sheet 6: 教师监考场次统计表 ----------


def _build_teacher_invigilation_stats_sheet(ws, exams, time_slots, teachers):
    """构建教师监考场次统计表（矩阵格式）

    格式与教务处模板一致：
    - 前8列为固定信息：序号、学院、教工号、姓名、联系电话、类别、本次承担任务、承担监考场次
    - 后续列为日期×时段矩阵，每天4个时段，有监考填1
    """
    # 只保留实际被考试使用的时段，避免数据库垃圾数据导致重复列
    used_slot_ids = set()
    for exam in exams:
        if exam.time_slot_id:
            used_slot_ids.add(exam.time_slot_id)

    # 过滤出被使用的时段，并按 (day_of_week, slot_code) 去重
    filtered_ts = [ts for ts in time_slots if ts.id in used_slot_ids]
    seen: set[tuple[int, str]] = set()
    unique_ts: list = []
    for ts in sorted(filtered_ts, key=lambda t: (t.day_of_week, t.slot_code, t.id)):
        key = (ts.day_of_week, ts.slot_code)
        if key not in seen:
            seen.add(key)
            unique_ts.append(ts)

    # 按 day_of_week 分组
    sorted_ts = sorted(unique_ts, key=lambda t: (t.day_of_week, t.slot_code))
    day_groups: dict[int, list] = {}
    for ts in sorted_ts:
        day_groups.setdefault(ts.day_of_week, []).append(ts)

    # 确保每天恰好4个时段（模板固定格式）
    # 如果某天时段不足，用占位符补齐以维持矩阵对齐
    fixed_cols = 8

    # ---------- 构建教师-时段映射 ----------
    teacher_slot_map: dict[int, set[int]] = {}
    for exam in exams:
        if exam.time_slot_id is None:
            continue
        for et in exam.teacher_assignments:
            teacher_slot_map.setdefault(et.teacher_id, set()).add(exam.time_slot_id)

    # 流动监考也计入
    teacher_map = {t.id: t for t in teachers}
    for ts in time_slots:
        for pt in ts.patrol_teachers:
            teacher_slot_map.setdefault(pt.teacher_id, set()).add(ts.id)

    # ---------- 表头第1行：日期行 ----------
    header_row1 = ["序号", "学院", "教工号", "姓名", "联系电话", "类别", "本次承担任务", "承担监考场次"]
    merge_ranges = []
    col_idx = fixed_cols + 1  # openpyxl 列号从1开始

    for day in range(1, 6):
        ts_list = day_groups.get(day, [])
        # 如果某天没有时段，也要占4列保持对齐
        span = max(len(ts_list), 4)
        # 日期显示：优先用 exam_date，否则用星期名称
        if ts_list and ts_list[0].exam_date:
            date_label = str(ts_list[0].exam_date)
        else:
            date_label = DAY_NAMES.get(day, f"周{day}")
        header_row1.append(date_label)
        # 后续补空占位以填满合并区域
        for _ in range(span - 1):
            header_row1.append("")
        if span > 1:
            merge_ranges.append((1, col_idx, 1, col_idx + span - 1))
        col_idx += span

    ws.append(header_row1)
    for cell in ws[1]:
        _set_header_style(cell)
    # 设置第1行高度
    ws.row_dimensions[1].height = 22

    # 执行合并单元格
    for start_row, start_col, end_row, end_col in merge_ranges:
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=end_row, end_column=end_col)

    # ---------- 表头第2行：时段行 ----------
    header_row2 = ["", "", "", "", "", "", "", ""]
    for day in range(1, 6):
        ts_list = day_groups.get(day, [])
        span = max(len(ts_list), 4)
        for i in range(span):
            if i < len(ts_list):
                # 提取时段编号数字：T1->1, T2->2, ...
                sc = ts_list[i].slot_code
                num = sc.replace("T", "").replace("t", "") if sc else str(i + 1)
                header_row2.append(num)
            else:
                header_row2.append(str(i + 1))

    ws.append(header_row2)
    for cell in ws[2]:
        _set_header_style(cell)
    ws.row_dimensions[2].height = 22

    # ---------- 数据行 ----------
    sorted_teachers = sorted(teachers, key=lambda t: t.id)
    for seq, teacher in enumerate(sorted_teachers, 1):
        slots = teacher_slot_map.get(teacher.id, set())
        type_str = "专职" if teacher.teacher_type.value == "full_time" else "兼职"

        row_data = [
            seq,       # 序号
            "",        # 学院
            "",        # 教工号
            teacher.name,
            "",        # 联系电话
            type_str,
            "",        # 本次承担任务（留空）
            len(slots),  # 承担监考场次
        ]

        for day in range(1, 6):
            ts_list = day_groups.get(day, [])
            span = max(len(ts_list), 4)
            for i in range(span):
                if i < len(ts_list):
                    ts = ts_list[i]
                    row_data.append(1 if ts.id in slots else "")
                else:
                    row_data.append("")

        ws.append(row_data)
        row_num = ws.max_row
        for cell in ws[row_num]:
            _set_cell_style(cell)
            # 序号、监考场次居中
            if cell.column == 1 or cell.column == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # 姓名单元格左对齐
            if cell.column == 4:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # ---------- 列宽调整 ----------
    # 固定列
    col_widths = [6, 12, 10, 10, 12, 8, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 动态列统一宽度
    for col in range(fixed_cols + 1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 5.5

    # 冻结窗格：冻结前8列和前2行
    ws.freeze_panes = "I3"


# ============================================================
# JSON 导出
# ============================================================


async def export_json(db: AsyncSession, version_id: int | None = None) -> dict[str, Any]:
    """导出排考结果为 JSON 格式

    Args:
        version_id: 可选，指定版本ID，仅导出版本对应的排考数据
    """
    exams = await _load_exams_with_relations(db, version_id=version_id)
    time_slots = await _load_time_slots(db)
    ts_map = {ts.id: ts for ts in time_slots}

    exam_list = []
    for exam in exams:
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        exam_data = {
            "id": exam.id,
            "course_id": exam.course_id,
            "course_name": exam.course.name,
            "course_type": exam.course.course_type.value,
            "exam_label": exam.exam_label.value if exam.exam_label else None,
            "status": exam.status.value,
            "is_locked": exam.is_locked,
            "time_slot": {
                "id": ts.id,
                "day_of_week": ts.day_of_week,
                "slot_code": ts.slot_code,
                "start_time": ts.start_time,
                "end_time": ts.end_time,
            } if ts else None,
            "classrooms": [
                {
                    "classroom_id": ec.classroom_id,
                    "classroom_name": ec.classroom.name if hasattr(ec, "classroom") and ec.classroom else None,
                    "total_students": ec.total_students,
                    "classes": [
                        {"class_id": ca.class_id, "student_count": ca.student_count}
                        for ca in ec.class_assignments
                    ],
                }
                for ec in exam.classroom_assignments
            ],
            "teachers": [
                {
                    "teacher_id": et.teacher_id,
                    "teacher_name": et.teacher.name if hasattr(et, "teacher") and et.teacher else None,
                    "role": et.role.value,
                    "classroom_id": et.classroom_id,
                }
                for et in exam.teacher_assignments
            ],
        }
        exam_list.append(exam_data)

    return {
        "export_time": datetime.now(CN_TZ).isoformat(),
        "total_exams": len(exam_list),
        "exams": exam_list,
    }


# ============================================================
# SQL 导出
# ============================================================


async def export_sql(db: AsyncSession, version_id: int | None = None) -> str:
    """导出排考结果为 SQL INSERT 语句

    Args:
        version_id: 可选，指定版本ID，仅导出版本对应的排考数据
    """
    exams = await _load_exams_with_relations(db, version_id=version_id)

    lines: list[str] = [
        "-- 考试排考系统 - 排考结果 SQL 导出",
        f"-- 导出时间: {datetime.now(CN_TZ).isoformat()}",
        f"-- 版本ID: {version_id if version_id else '全部'}",
        "BEGIN;",
        "",
    ]

    for exam in exams:
        # 快照数据是字典格式，需要特殊处理
        if isinstance(exam, dict):
            label = f"'{exam.get('exam_label', '')}'" if exam.get('exam_label') else "NULL"
            lines.append(
                f"-- 快照数据: course_id={exam.get('course_id')}, time_slot_id={exam.get('time_slot_id')}"
            )
            continue

        label = f"'{exam.exam_label.value}'" if exam.exam_label else "NULL"
        lines.append(
            f"INSERT INTO exams (id, course_id, time_slot_id, exam_label, status, is_locked, created_at, updated_at) "
            f"VALUES ({exam.id}, {exam.course_id}, {exam.time_slot_id or 'NULL'}, {label}, "
            f"'{exam.status.value}', {str(exam.is_locked).lower()}, NOW(), NOW()) "
            f"ON CONFLICT (id) DO UPDATE SET time_slot_id = EXCLUDED.time_slot_id, status = EXCLUDED.status;"
        )

        for ec in exam.classroom_assignments:
            lines.append(
                f"INSERT INTO exam_classrooms (exam_id, classroom_id, total_students) "
                f"VALUES ({ec.exam_id}, {ec.classroom_id}, {ec.total_students}) "
                f"ON CONFLICT (exam_id, classroom_id) DO UPDATE SET total_students = EXCLUDED.total_students;"
            )

        for et in exam.teacher_assignments:
            role = et.role.value
            cid = et.classroom_id if et.classroom_id else "NULL"
            lines.append(
                f"INSERT INTO exam_teachers (exam_id, teacher_id, role, classroom_id) "
                f"VALUES ({et.exam_id}, {et.teacher_id}, '{role}', {cid}) "
                f"ON CONFLICT (exam_id, teacher_id, role) DO NOTHING;"
            )

    lines.extend(["", "COMMIT;"])
    return "\n".join(lines)
