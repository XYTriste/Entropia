"""
从数据库导出排考结果为 Excel 文件
"""
import os
import sys
sys.path.insert(0, "/app")

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker, selectinload

from app.models.class_ import Class
from app.models.classroom import Classroom
from app.models.course import Course
from app.models.exam import Exam
from app.models.exam_classroom import ExamClassroom
from app.models.exam_classroom_class import ExamClassroomClass
from app.models.exam_teacher import ExamTeacher
from app.models.patrol_teacher import PatrolTeacher
from app.models.teacher import Teacher
from app.models.time_slot import TimeSlot

# ============================================================
# 样式常量
# ============================================================
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DAY_NAMES = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五"}


def _set_header_style(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = BORDER


def _set_cell_style(cell):
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER


def _auto_width(worksheet, min_width=10, max_width=40):
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        adjusted_width = min(max(length + 2, min_width), max_width)
        worksheet.column_dimensions[col_letter].width = adjusted_width


def main():
    db_url = "postgresql://scheduler:scheduler@db:5432/exam_scheduler"
    engine = create_engine(db_url)
    Session = sessionmaker(bind=engine)
    session = Session()

    exams = session.execute(
        select(Exam)
        .options(
            selectinload(Exam.course),
            selectinload(Exam.time_slot),
            selectinload(Exam.classroom_assignments).selectinload(ExamClassroom.class_assignments).selectinload(ExamClassroomClass.class_),
            selectinload(Exam.classroom_assignments).selectinload(ExamClassroom.classroom),
            selectinload(Exam.teacher_assignments).selectinload(ExamTeacher.teacher),
        )
        .order_by(Exam.time_slot_id)
    ).scalars().all()

    time_slots = session.execute(
        select(TimeSlot)
        .options(selectinload(TimeSlot.patrol_teachers).selectinload(PatrolTeacher.teacher))
        .order_by(TimeSlot.id)
    ).scalars().all()

    teachers = session.execute(select(Teacher)).scalars().all()
    classrooms_data = session.execute(select(Classroom)).scalars().all()
    classes_data = session.execute(select(Class)).scalars().all()

    wb = Workbook()

    # Sheet 1: 排考总览表
    _build_overview_sheet(wb.active, exams, time_slots, classrooms_data, teachers)
    wb.active.title = "排考总览表"

    # Sheet 2: 教师监考表
    wb.create_sheet("教师监考表")
    _build_teacher_sheet(wb["教师监考表"], exams, time_slots, teachers)

    # Sheet 3: 班级通知表
    wb.create_sheet("班级通知表")
    _build_class_notice_sheet(wb["班级通知表"], exams, time_slots, classes_data, classrooms_data)

    # Sheet 4: 考场签到表
    wb.create_sheet("考场签到表")
    _build_classroom_sign_sheet(wb["考场签到表"], exams, time_slots, classrooms_data)

    # Sheet 5: 流动监考巡查表
    wb.create_sheet("流动监考巡查表")
    _build_patrol_sheet(wb["流动监考巡查表"], time_slots, teachers)

    output_dir = "/app/output_results"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "schedule_result.xlsx")
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")
    print(f"Total exams: {len(exams)}")
    print(f"Total exam_classrooms: {sum(len(e.classroom_assignments) for e in exams)}")
    print(f"Total exam_teachers: {sum(len(e.teacher_assignments) for e in exams)}")
    print(f"Total patrol teachers (unique slots): {len(time_slots)}")


def _build_overview_sheet(ws, exams, time_slots, classrooms, teachers):
    headers = ["序号", "课程名称", "AB卷", "星期", "时段", "时间", "教室", "监考教师", "班级", "人数"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    ts_map = {ts.id: ts for ts in time_slots}
    teacher_map = {t.id: t for t in teachers}
    room_map = {r.id: r for r in classrooms}

    row_idx = 1
    seq = 0
    prev_course_label = None

    for exam in sorted(exams, key=lambda e: (e.time_slot_id or 0, e.course.name)):
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        day_str = DAY_NAMES.get(ts.day_of_week, "") if ts else ""
        time_str = f"{ts.start_time}-{ts.end_time}" if ts else ""
        label_str = exam.exam_label.value if exam.exam_label else ""
        course_label = (exam.course.name, label_str)

        if course_label != prev_course_label:
            seq += 1
            prev_course_label = course_label

        room_teachers = {}
        exam_fixed_teachers = []
        for et in exam.teacher_assignments:
            if et.role.value == "fixed":
                t = teacher_map.get(et.teacher_id)
                if t:
                    if et.classroom_id:
                        room_teachers.setdefault(et.classroom_id, []).append(t.name)
                    else:
                        exam_fixed_teachers.append(t.name)

        for ec in exam.classroom_assignments:
            room = room_map.get(ec.classroom_id)
            room_name = room.name if room else f"教室{ec.classroom_id}"
            teachers_in_room = room_teachers.get(ec.classroom_id, [])
            if teachers_in_room:
                teacher_str = "、".join(teachers_in_room)
            elif exam_fixed_teachers:
                teacher_str = "、".join(exam_fixed_teachers)
            else:
                teacher_str = ""

            if ec.class_assignments:
                for ca in ec.class_assignments:
                    cls_name = ca.class_.name if hasattr(ca, "class_") and ca.class_ else ""
                    row_idx += 1
                    ws.append([
                        seq, exam.course.name, label_str, day_str,
                        ts.slot_code if ts else "", time_str, room_name,
                        teacher_str, cls_name, ca.student_count,
                    ])
                    for cell in ws[row_idx]:
                        _set_cell_style(cell)
            else:
                row_idx += 1
                ws.append([
                    seq, exam.course.name, label_str, day_str,
                    ts.slot_code if ts else "", time_str, room_name,
                    teacher_str, "", ec.total_students,
                ])
                for cell in ws[row_idx]:
                    _set_cell_style(cell)

    _auto_width(ws)


def _build_teacher_sheet(ws, exams, time_slots, teachers):
    headers = ["教师姓名", "教师类型", "星期", "时段", "时间", "课程", "AB卷", "监考角色", "教室"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    ts_map = {ts.id: ts for ts in time_slots}
    room_map = {}
    for exam in exams:
        for ec in exam.classroom_assignments:
            room_name = ec.classroom.name if hasattr(ec, "classroom") and ec.classroom else f"教室{ec.classroom_id}"
            room_map[ec.classroom_id] = room_name

    teacher_exams = {}
    for exam in exams:
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        for et in exam.teacher_assignments:
            if et.teacher_id not in teacher_exams:
                teacher_exams[et.teacher_id] = []
            info = {
                "day": DAY_NAMES.get(ts.day_of_week, "") if ts else "",
                "slot": ts.slot_code if ts else "",
                "time": f"{ts.start_time}-{ts.end_time}" if ts else "",
                "course": exam.course.name,
                "label": exam.exam_label.value if exam.exam_label else "",
                "role": "固定监考" if et.role.value == "fixed" else "流动监考",
                "room": "",
            }
            if et.role.value == "fixed" and et.classroom_id:
                info["room"] = room_map.get(et.classroom_id, f"教室{et.classroom_id}")
            teacher_exams[et.teacher_id].append(info)

    row_idx = 1
    for teacher in sorted(teachers, key=lambda t: t.name):
        if teacher.id not in teacher_exams:
            continue
        for info in sorted(teacher_exams[teacher.id], key=lambda x: (x["day"], x["slot"])):
            row_idx += 1
            type_str = "专任" if teacher.teacher_type.value == "full_time" else "兼职"
            ws.append([
                teacher.name, type_str, info["day"], info["slot"],
                info["time"], info["course"], info["label"], info["role"], info["room"],
            ])
            for cell in ws[row_idx]:
                _set_cell_style(cell)

    _auto_width(ws)


def _build_class_notice_sheet(ws, exams, time_slots, classes_data, classrooms_data):
    headers = ["班级名称", "年级", "星期", "时段", "时间", "课程", "AB卷", "考场", "考生数"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    ts_map = {ts.id: ts for ts in time_slots}
    class_map = {c.id: c for c in classes_data}
    room_map = {r.id: r for r in classrooms_data}

    row_idx = 1
    for exam in sorted(exams, key=lambda e: e.time_slot_id or 0):
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        for ec in exam.classroom_assignments:
            for ca in ec.class_assignments:
                cls = class_map.get(ca.class_id)
                if not cls:
                    continue
                room = room_map.get(ec.classroom_id)
                row_idx += 1
                ws.append([
                    cls.name, cls.grade,
                    DAY_NAMES.get(ts.day_of_week, "") if ts else "",
                    ts.slot_code if ts else "",
                    f"{ts.start_time}-{ts.end_time}" if ts else "",
                    exam.course.name,
                    exam.exam_label.value if exam.exam_label else "",
                    room.name if room else f"教室{ec.classroom_id}",
                    ca.student_count,
                ])
                for cell in ws[row_idx]:
                    _set_cell_style(cell)

    _auto_width(ws)


def _build_classroom_sign_sheet(ws, exams, time_slots, classrooms_data):
    headers = ["教室名称", "星期", "时段", "时间", "课程", "AB卷", "应到人数", "监考教师"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    ts_map = {ts.id: ts for ts in time_slots}
    teacher_map = {}
    for exam in exams:
        for et in exam.teacher_assignments:
            if hasattr(et, "teacher") and et.teacher:
                teacher_map[et.teacher_id] = et.teacher.name
    room_map = {r.id: r for r in classrooms_data}

    row_idx = 1
    for exam in sorted(exams, key=lambda e: e.time_slot_id or 0):
        ts = ts_map.get(exam.time_slot_id) if exam.time_slot_id else None
        for ec in exam.classroom_assignments:
            room = room_map.get(ec.classroom_id)
            fixed_t = []
            for et in exam.teacher_assignments:
                if et.role.value == "fixed" and et.classroom_id == ec.classroom_id:
                    name = teacher_map.get(et.teacher_id, f"教师{et.teacher_id}")
                    fixed_t.append(name)

            row_idx += 1
            ws.append([
                room.name if room else f"教室{ec.classroom_id}",
                DAY_NAMES.get(ts.day_of_week, "") if ts else "",
                ts.slot_code if ts else "",
                f"{ts.start_time}-{ts.end_time}" if ts else "",
                exam.course.name,
                exam.exam_label.value if exam.exam_label else "",
                ec.total_students,
                "; ".join(fixed_t),
            ])
            for cell in ws[row_idx]:
                _set_cell_style(cell)

    _auto_width(ws)


def _build_patrol_sheet(ws, time_slots, teachers):
    headers = ["星期", "时段", "时间", "流动监考1", "流动监考2", "流动监考3"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)

    teacher_map = {t.id: t for t in teachers}

    row_idx = 1
    for ts in sorted(time_slots, key=lambda t: (t.day_of_week, t.slot_code)):
        patrol_names = []
        for pt in ts.patrol_teachers:
            t = teacher_map.get(pt.teacher_id)
            patrol_names.append(t.name if t else f"教师{pt.teacher_id}")
        while len(patrol_names) < 3:
            patrol_names.append("")

        row_idx += 1
        ws.append([
            DAY_NAMES.get(ts.day_of_week, ""),
            ts.slot_code,
            f"{ts.start_time}-{ts.end_time}",
            patrol_names[0], patrol_names[1], patrol_names[2],
        ])
        for cell in ws[row_idx]:
            _set_cell_style(cell)

    _auto_width(ws)


if __name__ == "__main__":
    main()
